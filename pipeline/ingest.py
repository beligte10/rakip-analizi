"""
pipeline.ingest
================
BDDK xlsx dosyalarını long-format parquet'e dönüştürür.

İki kullanım:

1) Tek dosya yükle (Admin upload sonrası):
    from pipeline.ingest import ingest_files
    ingest_files(['/tmp/Akbank - 30.09.2025.xlsx'])

2) Tüm raw klasörü taranıp baştan parquet üret (init):
    from pipeline.ingest import rebuild_parquet
    rebuild_parquet('/path/to/data/raw', '/path/to/data/veriler.parquet')
"""
from __future__ import annotations
import os
import re
import json
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional

import pandas as pd
import openpyxl


FILENAME_RE = re.compile(r'^(.+?)\s*-\s*(\d{2})\.(\d{2})\.(\d{4})\.xlsx?$', re.IGNORECASE)

# Bir dosyada bundan fazla '#VALUE!' hücresi varsa dosya Bloomberg/FactSet
# eklentisi olmadan export edilmiş sayılır (bkz. check_data_quality).
# Eşik, 2026-08-11'de bulunan gerçek bozuk dosyalarda (QNB Finansbank vb.
# ~2400-2600 hücre) ile normal/erken-yıl şablon farklarında (dipnot kalemleri
# BOŞ '' olur, '#VALUE!' DEĞİL — en fazla birkaç yüz boş hücre görülür, hiç
# '#VALUE!' görülmez) net bir ayrım sağlıyor.
VALUE_ERROR_THRESHOLD = 50


# ============================================================
# Tek xlsx → DataFrame
# ============================================================

def parse_filename(filename: str) -> Optional[tuple[str, str]]:
    """'<Banka> - DD.MM.YYYY.xlsx' → (banka, 'YYYY-MM-DD')"""
    m = FILENAME_RE.match(filename.strip())
    if not m:
        return None
    banka = m.group(1).strip()
    d, mo, y = m.group(2), m.group(3), m.group(4)
    return banka, f'{y}-{mo}-{d}'


def load_xlsx(path: Path | str, banka: str, tarih: str,
              banka_turu: Optional[str] = None) -> pd.DataFrame:
    """
    Tek bir BDDK xlsx'ini long-format DataFrame'e çevir.

    Beklenen yapı:
    - Sheet adı: 'Sheet1'
    - 13. satır header (`Banka Türü, Tablo Türü, Tablo Adı, Kalem Adı,
                        Para Birimi, Item Code, Tutar`)
    - 14. satırdan itibaren data
    """
    path = Path(path)
    df = pd.read_excel(path, sheet_name='Sheet1', header=13, engine='openpyxl')

    expected_cols = ['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı',
                     'Para Birimi', 'Tutar']
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        raise ValueError(f"{path.name}: eksik sütunlar: {missing}")

    df = df[['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı',
             'Para Birimi', 'Tutar']].copy()
    df['Tutar'] = pd.to_numeric(df['Tutar'], errors='coerce')
    df = df.dropna(subset=['Tutar', 'Kalem Adı'])

    df['Banka Adı'] = banka
    df['Tarih'] = pd.Timestamp(tarih)

    if banka_turu and df['Banka Türü'].isna().any():
        df['Banka Türü'] = df['Banka Türü'].fillna(banka_turu)

    return df


def load_dir(banka_dir: Path, banka: str,
             banka_turu: Optional[str] = None) -> pd.DataFrame:
    """Bir banka klasöründeki tüm xlsx'leri yükle."""
    frames = []
    for f in sorted(banka_dir.glob('*.xlsx')):
        parsed = parse_filename(f.name)
        if not parsed:
            continue
        _, tarih = parsed
        try:
            frames.append(load_xlsx(f, banka, tarih, banka_turu))
        except Exception as e:
            print(f"  ! Skip {f.name}: {e}")
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def _load_one_file_worker(args: tuple[str, str, str, Optional[str]]):
    """ProcessPoolExecutor worker — modül seviyesinde olmalı (pickle için).
    Hata durumunda None döner (tek dosyanın bozukluğu tüm rebuild'i düşürmesin)."""
    path_str, banka, tarih, banka_turu = args
    try:
        return banka, load_xlsx(Path(path_str), banka, tarih, banka_turu)
    except Exception as e:
        print(f"  ! Skip {Path(path_str).name}: {e}")
        return banka, None


# ============================================================
# Validation
# ============================================================

def validate_filename(filename: str, valid_banks: List[str]) -> tuple[bool, str]:
    """Dosya adı kuralına uyuyor mu, banka tanımlı mı?"""
    parsed = parse_filename(filename)
    if not parsed:
        return False, f"Dosya adı '<Banka> - DD.MM.YYYY.xlsx' formatında olmalı"
    banka, tarih = parsed
    if banka not in valid_banks:
        return False, f"Banka '{banka}' catalog'ta tanımlı değil"
    # Q-end kontrolü
    try:
        d = datetime.strptime(tarih, '%Y-%m-%d').date()
        if (d.month, d.day) not in [(3, 31), (6, 30), (9, 30), (12, 31)]:
            return False, f"Tarih ({tarih}) bir çeyrek sonu değil"
    except ValueError:
        return False, f"Tarih ({tarih}) geçersiz"
    return True, ""


def check_data_quality(source) -> tuple[bool, str]:
    """
    xlsx içeriği sağlıklı mı? (`source`: dosya yolu veya dosya-benzeri nesne/BytesIO)

    Bloomberg/FactSet eklentisi olmadan export edilen dosyalarda Tutar
    hücreleri '#VALUE!' hata metniyle ya da boş string ile dolar. Bu fark
    edilmezse pipeline/lookup.py::LookupContext._lookup() eksik veriyi
    (KeyError) sessizce 0.0'a çevirdiğinden, ölçütler "gerçek sıfır" gibi
    görünür — hiçbir hata/uyarı üretmeden (bkz. 2026-08-11: QNB Finansbank
    2013-2024 arası ~42 dosya, Kuveyt Türk 2013-2014 arası 4 dosya bu şekilde
    bozuktu, elle bulunup düzeltildi).

    Kontrol:
    1. 'Toplam Aktifler' (Bilanço, Ana Tablo, Toplam) hücresi sayısal mı?
       (Her banka/çeyrek için var olması gereken, en temel bilanço kalemi.)
    2. Dosyada VALUE_ERROR_THRESHOLD'dan fazla '#VALUE!' hücresi var mı?
    """
    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except Exception as e:
        return False, f"Dosya okunamadı/bozuk: {e}"

    if 'Sheet1' not in wb.sheetnames:
        return False, "'Sheet1' sayfası bulunamadı"

    ws = wb['Sheet1']
    ta = None
    value_err = 0
    for row in ws.iter_rows(min_row=15, max_col=7, values_only=True):
        if len(row) < 7:
            continue
        if row[6] == '#VALUE!':
            value_err += 1
        if row[2] == 'Bilanço' and row[3] == 'Toplam Aktifler' and row[4] == 'Toplam':
            ta = row[6]

    if not isinstance(ta, (int, float)):
        return False, (
            "'Toplam Aktifler' hücresi boş/geçersiz — dosya muhtemelen "
            "Bloomberg/FactSet eklentisi olmadan export edilmiş"
        )
    if value_err > VALUE_ERROR_THRESHOLD:
        return False, (
            f"{value_err} hücrede '#VALUE!' hatası — dosya muhtemelen "
            "Bloomberg/FactSet eklentisi olmadan export edilmiş"
        )
    return True, ""


# ============================================================
# Parquet rebuild & incremental
# ============================================================

def rebuild_parquet(raw_dir: Path | str, parquet_path: Path | str,
                     bank_turu_map: Optional[dict] = None,
                     max_workers: Optional[int] = None) -> pd.DataFrame:
    """
    raw_dir altındaki tüm banka klasörlerini tarayıp tek parquet üretir.
    Her banka klasörü = bir display name olmalı.

    Dosya okuma (openpyxl ile xlsx→DataFrame, CPU-yoğun bir iş) dosya
    seviyesinde PARALEL çalışır — 1170 dosyalık tam arşivde tek çekirdekte
    ~70 sn süren bu adım, çok çekirdekli makinelerde birkaç saniyeye iner.
    max_workers=None ise os.cpu_count() kullanılır; tek dosyalık/az dosyalı
    incremental upload'larda paralellik overhead'i anlamsız olduğundan
    max_workers=1 verilirse sıralı (eski) davranışa döner.
    """
    raw_dir = Path(raw_dir)
    if bank_turu_map is None:
        bank_turu_map = {}

    # Tüm banka klasörlerindeki xlsx'leri tek bir görev listesine topla
    tasks: list[tuple[str, str, str, Optional[str]]] = []
    for banka_dir in sorted(raw_dir.iterdir()):
        if not banka_dir.is_dir():
            continue
        banka = banka_dir.name
        bt = bank_turu_map.get(banka)
        for f in sorted(banka_dir.glob('*.xlsx')):
            parsed = parse_filename(f.name)
            if not parsed:
                continue
            _, tarih = parsed
            tasks.append((str(f), banka, tarih, bt))

    if not tasks:
        raise RuntimeError(f"Hiç veri yüklenemedi: {raw_dir}")

    by_bank: dict[str, list[pd.DataFrame]] = {}
    workers = max_workers if max_workers is not None else min(os.cpu_count() or 1, 10)
    if workers <= 1 or len(tasks) <= 1:
        results = (_load_one_file_worker(t) for t in tasks)
    else:
        with ProcessPoolExecutor(max_workers=workers) as ex:
            results = list(ex.map(_load_one_file_worker, tasks))

    for banka, df in results:
        if df is not None and len(df):
            by_bank.setdefault(banka, []).append(df)

    if not by_bank:
        raise RuntimeError(f"Hiç veri yüklenemedi: {raw_dir}")

    frames = []
    for banka, dfs in sorted(by_bank.items()):
        df = pd.concat(dfs, ignore_index=True)
        frames.append(df)
        print(f"  {banka}: {len(df):,} satır, {df['Tarih'].nunique()} dönem")

    full = pd.concat(frames, ignore_index=True)
    for col in ['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı', 'Para Birimi', 'Banka Adı']:
        full[col] = full[col].astype('category')

    parquet_path = Path(parquet_path)
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    full.to_parquet(parquet_path, index=False, compression='zstd')
    print(f"\n✓ Parquet: {parquet_path} ({parquet_path.stat().st_size / 1024 / 1024:.1f} MB)")
    return full


def update_parquet_incremental(
    new_files: List[tuple[Path | str, str, str, Optional[str]]],
    parquet_path: Path | str,
) -> pd.DataFrame:
    """
    Mevcut parquet'i BAŞTAN taramadan sadece verilen dosyaları ekler/günceller.

    `new_files`: (path, banka, tarih, banka_turu) tuple listesi.

    `rebuild_parquet`'ten farkı: data/raw/'daki TÜM dosyaları değil, SADECE
    verilenleri parse eder. `admin_upload` (tek/az sayıda dosyalık çeyreklik
    yükleme) için tasarlandı — 1176 dosyalık tam arşivde ~21s süren xlsx
    okuma adımını, N (genelde 1-27) dosyalık bir yükleme için ~1-2s'ye
    indirir (bkz. 2026-08-11 verimlilik raporu, madde 2). `admin_rebuild` /
    `admin_upload_zip` gibi "her şeyi yeniden tara" senaryoları hâlâ
    `rebuild_parquet` kullanmalı — onlarda zaten TÜM dosyalar okunacağından
    incremental'ın bir faydası olmaz.

    Üzerine yazma (aynı banka+tarih tekrar yüklenirse): mevcut parquet'teki
    o (Banka Adı, Tarih) çiftine ait ESKİ satırlar çıkarılır, yenileri
    eklenir — filtreleme pandas `merge` ile VEKTÖRİZE yapılır (parquet
    milyonlarca satır olabileceğinden row-wise `.apply` kullanılmaz).

    ProcessPoolExecutor KULLANILMAZ — az sayıda dosyada process spin-up
    overhead'i (her worker ~0.1-0.3s) faydadan fazla olurdu; bu zaten
    `rebuild_parquet`'in de `len(tasks) <= 1` durumunda sıralı çalışma
    mantığıyla tutarlı.
    """
    parquet_path = Path(parquet_path)
    cols = ['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı',
            'Para Birimi', 'Tutar', 'Banka Adı', 'Tarih']

    if parquet_path.exists():
        existing = pd.read_parquet(parquet_path)
        for col in ['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı',
                    'Para Birimi', 'Banka Adı']:
            if col in existing.columns:
                existing[col] = existing[col].astype(str)
    else:
        existing = pd.DataFrame(columns=cols)

    new_frames = []
    touched: set[tuple[str, pd.Timestamp]] = set()
    for path, banka, tarih, banka_turu in new_files:
        df = load_xlsx(path, banka, tarih, banka_turu)
        new_frames.append(df)
        touched.add((banka, pd.Timestamp(tarih)))

    if not new_frames:
        return existing

    new_df = pd.concat(new_frames, ignore_index=True)

    if len(existing) and touched:
        touched_banks = {b for b, _ in touched}
        is_candidate = existing['Banka Adı'].isin(touched_banks)
        rest = existing[~is_candidate]
        candidates = existing[is_candidate]

        touched_df = pd.DataFrame(list(touched), columns=['Banka Adı', 'Tarih'])
        merged = candidates.merge(
            touched_df, on=['Banka Adı', 'Tarih'], how='left', indicator=True
        )
        keep_mask = (merged['_merge'] == 'left_only').to_numpy()
        existing_filtered = pd.concat(
            [rest, candidates.loc[keep_mask]], ignore_index=True
        )
    else:
        existing_filtered = existing

    full = pd.concat([existing_filtered, new_df], ignore_index=True)
    for col in ['Banka Türü', 'Tablo Türü', 'Tablo Adı', 'Kalem Adı', 'Para Birimi', 'Banka Adı']:
        full[col] = full[col].astype('category')

    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = parquet_path.with_suffix('.tmp')
    full.to_parquet(tmp, index=False, compression='zstd')
    tmp.replace(parquet_path)
    print(f"✓ Parquet (incremental): {len(new_files)} dosya işlendi, "
          f"toplam {len(full):,} satır ({parquet_path.stat().st_size / 1024 / 1024:.1f} MB)")
    return full


def ingest_files(file_paths: Iterable[Path | str],
                  raw_dir: Path | str,
                  parquet_path: Path | str,
                  bank_turu_map: Optional[dict] = None) -> List[str]:
    """
    Yüklenen xlsx'leri raw_dir'a kopyalar ve parquet'i incremental günceller
    (bkz. update_parquet_incremental — data/raw/'daki diğer dosyalar
    yeniden okunmaz).

    Dönüş: işlenen dosyaların isim listesi.
    """
    raw_dir = Path(raw_dir)
    raw_dir.mkdir(parents=True, exist_ok=True)
    if bank_turu_map is None:
        bank_turu_map = {}

    processed = []
    to_ingest: list[tuple[Path, str, str, Optional[str]]] = []
    for fp in file_paths:
        fp = Path(fp)
        parsed = parse_filename(fp.name)
        if not parsed:
            print(f"  ! Skip {fp.name}: dosya adı uygunsuz")
            continue
        banka, tarih = parsed
        target_dir = raw_dir / banka
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / fp.name
        # Dosyayı kopyala
        target.write_bytes(fp.read_bytes())
        processed.append(fp.name)
        to_ingest.append((target, banka, tarih, bank_turu_map.get(banka)))

    if to_ingest:
        update_parquet_incremental(to_ingest, parquet_path)

    return processed
